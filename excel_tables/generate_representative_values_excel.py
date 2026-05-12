from __future__ import annotations

import argparse
import csv
from copy import copy
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


ROOT_DIR = Path(__file__).resolve().parents[1]
PROCESSED_DIR = ROOT_DIR / "processed"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "tablas_valores_representativos_para_docx_acentos_corregidos.xlsx"
TABLE_FONT_SIZE = 8
TITLE_FONT_SIZE = 9

SAMPLE_TYPE_ES = {
    "Fresh manure": "Estiércol fresco",
    "Precomposted manure": "Estiércol precompostado",
}

TREATMENT_LABELS = {
    "A": "A - estiércol fresco",
    "B": "B - estiércol precompostado",
}

SAMPLE_LABELS = {
    "A1": "Estiércol fresco 1",
    "A2": "Estiércol fresco 2",
    "B1": "Estiércol precompostado 1",
    "B2": "Estiércol precompostado 2",
}

N_TREATMENT_ES = {
    "ESTIERCOL FRESCO": "Estiércol fresco",
    "LIQ: AGUA VERDE": "Líquido: agua verde",
    "LIQ: PURINES": "Líquido: purines",
    "SOL: PRECOMPOSTADO": "Sólido precompostado",
}


def read_csv(name: str) -> list[dict[str, str]]:
    path = PROCESSED_DIR / name
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def fmt_date(value: str) -> str:
    value = value.strip()
    if not value:
        return ""
    for pattern in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, pattern).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return value


def number(value: str, digits: int = 3) -> float | int | str:
    value = value.strip()
    if not value:
        return ""
    parsed = round(float(value), digits)
    return int(parsed) if parsed.is_integer() else parsed


def add_rows(ws, rows: list[list[Any]], title_rows: set[int] | None = None, header_rows: set[int] | None = None) -> None:
    title_rows = title_rows or set()
    header_rows = header_rows or set()
    horizontal_line = Side(style="thin", color="000000")
    row_border = Border(bottom=horizontal_line)
    max_columns = max(len(row) for row in rows)

    for row_index, row in enumerate(rows, start=1):
        ws.append(row)
        if not any(value not in (None, "") for value in row):
            continue
        for column_index in range(1, max_columns + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = row_border
            cell.font = Font(size=TABLE_FONT_SIZE)
            if row_index in title_rows:
                cell.font = Font(bold=True, size=TITLE_FONT_SIZE)
            if row_index in header_rows:
                cell.font = Font(bold=True, size=TABLE_FONT_SIZE)


def set_column_widths(ws, widths: list[float]) -> None:
    for column_index, width in enumerate(widths, start=1):
        column_letter = ws.cell(row=1, column=column_index).column_letter
        ws.column_dimensions[column_letter].width = width


def configure_letter_page(ws) -> None:
    ws.page_setup.paperSize = 1
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def enable_multiline_cells(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.horizontal = "center"
                alignment.vertical = "center"
                cell.alignment = alignment


def build_workbook() -> Workbook:
    solids_by_sample = read_csv("volatile_solids_representative_table.csv")
    solids_by_treatment = read_csv("volatile_solids_treatment_table.csv")
    nitrogen_samples = read_csv("CIA_samples_table_v6.csv")
    nitrogen_summary = read_csv("CIA_samples_table_v6_treatment_summary.csv")

    nitrogen_by_treatment = {row["treatment"]: row for row in nitrogen_summary}

    wb = Workbook()
    ws_docx = wb.active
    ws_docx.title = "Tablas para docx"

    table_1_header = [
        "Muestra",
        "Tipo de muestra",
        "Fecha de análisis de sólidos",
        "Número de muestras de sólidos (n)",
        "Contenido de masa seca promedio (% base húmeda)",
        "Desviación estándar de masa seca (%)",
        "Sólidos volátiles promedio (% base seca)",
        "Desviación estándar de sólidos volátiles (%)",
        "Fecha de análisis de nitrógeno",
        "Número de muestras de nitrógeno (n)",
        "Nitrógeno total promedio (% masa)",
        "Nitrógeno total promedio (mg/kg)",
    ]
    nitrogen_lookup = {
        "A": nitrogen_by_treatment["ESTIERCOL FRESCO"],
        "B": nitrogen_by_treatment["SOL: PRECOMPOSTADO"],
    }
    table_1 = [["Tabla 1. Valores representativos por tipo de muestra sólida"], table_1_header]
    for row in solids_by_treatment:
        treatment = row["treatment"]
        n_row = nitrogen_lookup[treatment]
        table_1.append(
            [
                TREATMENT_LABELS[treatment],
                SAMPLE_TYPE_ES[row["sample_type"]],
                fmt_date(row["sampling_date"]),
                number(row["sample_count"], 0),
                number(row["dry_matter_treatment_mean_pct"]),
                number(row["dry_matter_treatment_sd_pct"]),
                number(row["volatile_solids_treatment_mean_pct"]),
                number(row["volatile_solids_treatment_sd_pct"]),
                fmt_date(n_row["date"]),
                number(n_row["n_samples"], 0),
                number(n_row["mean_n_percentage"]),
                number(n_row["mean_n_total_mg_kg"]),
            ]
        )

    table_2_header = [
        "Muestra",
        "Fecha de análisis",
        "Número de réplicas (n)",
        "Contenido de masa seca promedio (% base húmeda)",
        "Desviación estándar de masa seca (%)",
        "Coeficiente de variación de masa seca (%)",
        "Sólidos volátiles promedio (% base seca)",
        "Desviación estándar de sólidos volátiles (%)",
        "Coeficiente de variación de sólidos volátiles (%)",
    ]
    table_2 = [[], [], ["Tabla 2. Contenido de masa seca y sólidos volátiles por muestra"], table_2_header]
    for row in solids_by_sample:
        sample_type = SAMPLE_TYPE_ES[row["sample_type"]]
        table_2.append(
            [
                SAMPLE_LABELS.get(row["sample_group"], f"{row['sample_group']} - {sample_type}"),
                fmt_date(row["sampling_date"]),
                number(row["replicate_count"], 0),
                number(row["dry_matter_mean_pct"]),
                number(row["dry_matter_sd_pct"]),
                number(row["dry_matter_cv_pct"]),
                number(row["volatile_solids_mean_pct"]),
                number(row["volatile_solids_sd_pct"]),
                number(row["volatile_solids_cv_pct"]),
            ]
        )

    table_3_header = [
        "Muestra o tratamiento",
        "Fecha de análisis",
        "Número de muestras (n)",
        "Nitrógeno total promedio (% masa)",
        "Nitrógeno total promedio (mg/kg)",
        "Mediana de nitrógeno total (% masa)",
        "Mediana de nitrógeno total (mg/kg)",
        "Valor mínimo de nitrógeno total (mg/kg)",
        "Valor máximo de nitrógeno total (mg/kg)",
    ]
    table_3 = [[], [], ["Tabla 3. Nitrógeno total representativo por muestra o tratamiento"], table_3_header]
    for row in nitrogen_summary:
        table_3.append(nitrogen_summary_row(row))

    add_rows(
        ws_docx,
        table_1 + table_2 + table_3,
        title_rows={1, 7, 15},
        header_rows={2, 8, 16},
    )
    set_column_widths(ws_docx, [9, 9, 7, 6.5, 8, 7.5, 8, 7.5, 7, 6.5, 6, 6])
    configure_letter_page(ws_docx)

    ws_solids = wb.create_sheet("Sólidos y masa seca")
    solids_rows = [["Contenido de masa seca y sólidos volátiles por muestra"]]
    solids_rows.append(
        [
            "Muestra",
            "Tipo de muestra",
            "Fecha de análisis",
            "Número de réplicas (n)",
            "Contenido de masa seca promedio (% base húmeda)",
            "Desviación estándar de masa seca (%)",
            "Coeficiente de variación de masa seca (%)",
            "Sólidos volátiles promedio (% base seca)",
            "Desviación estándar de sólidos volátiles (%)",
            "Coeficiente de variación de sólidos volátiles (%)",
        ]
    )
    for row in solids_by_sample:
        solids_rows.append(
            [
                SAMPLE_LABELS.get(row["sample_group"], row["sample_group"]),
                SAMPLE_TYPE_ES[row["sample_type"]],
                fmt_date(row["sampling_date"]),
                number(row["replicate_count"], 0),
                number(row["dry_matter_mean_pct"]),
                number(row["dry_matter_sd_pct"]),
                number(row["dry_matter_cv_pct"]),
                number(row["volatile_solids_mean_pct"]),
                number(row["volatile_solids_sd_pct"]),
                number(row["volatile_solids_cv_pct"]),
            ]
        )
    add_rows(ws_solids, solids_rows, title_rows={1}, header_rows={2})
    set_column_widths(ws_solids, [7, 9, 7, 6, 8, 7, 7, 8, 7, 7])
    configure_letter_page(ws_solids)

    ws_n = wb.create_sheet("Nitrógeno total")
    n_rows = [["Nitrógeno total por muestra individual"]]
    n_rows.append(["Muestra", "Tipo de análisis", "Fecha de análisis", "Nitrógeno total (% masa)", "Nitrógeno total (mg/kg)"])
    for row in nitrogen_samples:
        n_rows.append(
            [
                row["sample_id"],
                row["analysis_type"],
                fmt_date(row["date"]),
                number(row["n_total_porcentaje"]),
                number(row["n_total_mg_kg"]),
            ]
        )
    n_rows.extend([[], [], ["Nitrógeno total representativo por muestra o tratamiento"], table_3_header])
    for row in nitrogen_summary:
        n_rows.append(nitrogen_summary_row(row))
    add_rows(ws_n, n_rows, title_rows={1, 13}, header_rows={2, 14})
    set_column_widths(ws_n, [11, 10, 7, 7, 8, 7, 8, 7, 7])
    configure_letter_page(ws_n)

    ws_notes = wb.create_sheet("Notas")
    add_rows(
        ws_notes,
        [
            ["Criterio de cálculo"],
            [
                "Los valores de contenido de masa seca y sólidos volátiles provienen de "
                "processed/volatile_solids_representative_table.csv y processed/volatile_solids_treatment_table.csv."
            ],
            [
                "El contenido de masa seca representativo por muestra corresponde al promedio de las réplicas "
                "agrupadas por muestra de estiércol fresco o estiércol precompostado."
            ],
            ["Los sólidos volátiles se calcularon como 100 menos el porcentaje de cenizas, en base seca."],
            [
                "El nitrógeno total representativo proviene de processed/CIA_samples_table_v6_treatment_summary.csv "
                "y corresponde al promedio por tratamiento usado por el script."
            ],
            ["Las unidades están indicadas en cada encabezado para facilitar la copia directa al documento académico."],
        ],
        title_rows={1},
    )
    ws_notes.column_dimensions["A"].width = 85
    configure_letter_page(ws_notes)
    for row in ws_notes.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")

    enable_multiline_cells(wb)
    return wb


def nitrogen_summary_row(row: dict[str, str]) -> list[Any]:
    return [
        N_TREATMENT_ES[row["treatment"]],
        fmt_date(row["date"]),
        number(row["n_samples"], 0),
        number(row["mean_n_percentage"]),
        number(row["mean_n_total_mg_kg"]),
        number(row["n_total_pct_median"]),
        number(row["n_total_mg_kg_median"]),
        number(row["n_total_mg_kg_min"]),
        number(row["n_total_mg_kg_max"]),
    ]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera el Excel de tablas de valores representativos con acentos corregidos."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Ruta del archivo .xlsx de salida. Valor por defecto: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output if args.output.is_absolute() else ROOT_DIR / args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
