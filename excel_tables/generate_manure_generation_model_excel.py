from __future__ import annotations

import argparse
import csv
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


ROOT_DIR = Path(__file__).resolve().parents[1]
PROCESSED_DIR = ROOT_DIR / "processed"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "tabla_generacion_estiercol_modelo.xlsx"
TABLE_FONT_SIZE = 8
TITLE_FONT_SIZE = 9

STAGE_LABELS = {
    ("A", "1"): "Precomposteo",
    ("A", "2"): "Lombricompostaje",
    ("A", "3"): "Almacenamiento de aguas verdes",
    ("A", "4"): "Aplicación de aguas verdes en campo",
    ("B", "1"): "Almacenamiento de purines",
    ("B", "2"): "Aplicación directa de purines en campo",
}

SCENARIO_LABELS = {
    "A": "A - lombricompostaje y aguas verdes",
    "B": "B - aplicación directa de purines",
}


def read_csv(name: str) -> list[dict[str, str]]:
    path = PROCESSED_DIR / name
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def number(value: str | float | int, digits: int = 3) -> float | int | str:
    if value in (None, ""):
        return ""
    parsed = round(float(value), digits)
    return int(parsed) if parsed.is_integer() else parsed


def decimal(value: str | float | int, digits: int = 3) -> str:
    if value in (None, ""):
        return ""
    return f"{float(value):.{digits}f}"


def add_rows(
    ws,
    rows: list[list[Any]],
    title_rows: set[int] | None = None,
    header_rows: set[int] | None = None,
) -> None:
    title_rows = title_rows or set()
    header_rows = header_rows or set()
    horizontal_line = Side(style="thin", color="000000")
    row_border = Border(bottom=horizontal_line)
    max_columns = max(len(row) for row in rows)

    for row_index, row in enumerate(rows, start=1):
        ws.append(row)
        if not any(value not in (None, "") for value in row):
            continue
        for column_index in range(1, max_columns + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = row_border
            cell.font = Font(size=TABLE_FONT_SIZE)
            if row_index in title_rows:
                cell.font = Font(bold=True, size=TITLE_FONT_SIZE)
            if row_index in header_rows:
                cell.font = Font(bold=True, size=TABLE_FONT_SIZE)


def set_column_widths(ws, widths: list[float]) -> None:
    for column_index, width in enumerate(widths, start=1):
        column_letter = ws.cell(row=1, column=column_index).column_letter
        ws.column_dimensions[column_letter].width = width


def configure_letter_page(ws) -> None:
    ws.page_setup.paperSize = 1
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def enable_multiline_cells(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.horizontal = "center"
                alignment.vertical = "center"
                cell.alignment = alignment


def build_base_generation_rows(stats: list[dict[str, str]]) -> list[list[Any]]:
    manure = next(row for row in stats if row["variable"].strip().lower() == "boniga")
    return [
        ["Tabla 1. Generación de estiércol fresco utilizada como entrada del modelo"],
        [
            "Variable",
            "Unidad",
            "n",
            "Duración del muestreo (días)",
            "Promedio por muestreo",
            "Mediana",
            "Mínimo",
            "Máximo",
            "Desviación estándar",
            "Flujo diario",
            "Flujo semanal",
            "Flujo anual usado en el modelo",
        ],
        [
            "Boñiga fresca recolectada",
            manure["unidad"],
            number(manure["n_datos"], 0),
            decimal(manure["duracion_muestreo_dias"], 1),
            decimal(manure["promedio"], 3),
            decimal(manure["mediana"], 3),
            decimal(manure["minimo"], 3),
            decimal(manure["maximo"], 3),
            decimal(manure["desviacion_estandar"], 3),
            decimal(manure["flujo_por_dia"], 3),
            decimal(manure["flujo_por_semana"], 3),
            decimal(manure["flujo_por_ano"], 3),
        ],
    ]


def build_model_mass_rows(mass_rows: list[dict[str, str]]) -> list[list[Any]]:
    rows = [
        [],
        [],
        ["Tabla 2. Masa anual de estiércol fresco asignada a cada escenario y etapa del modelo"],
        [
            "Escenario",
            "Etapa",
            "Proceso representado",
            "Factor de asignación de boñiga",
            "Boñiga fresca usada por el modelo (kg año⁻¹)",
            "Agua incorporada (L año⁻¹)",
            "Masa total equivalente (kg año⁻¹)",
            "Origen del valor",
        ],
    ]
    for row in mass_rows:
        escenario = row["escenario"]
        etapa = row["etapa"]
        rows.append(
            [
                SCENARIO_LABELS.get(escenario, escenario),
                number(etapa, 0),
                STAGE_LABELS.get((escenario, etapa), ""),
                decimal(row["factor_boniga_override"], 2),
                decimal(row["boniga_kg"], 3),
                decimal(row["agua_l"], 3),
                decimal(row["masa_total_kg_eq"], 3),
                "Factor de asignación del modelo",
            ]
        )
    return rows


def build_workbook() -> Workbook:
    stats = read_csv("agua_boniga_estadistica_descriptiva.csv")
    mass_rows = read_csv("masa_total_escenario_etapa.csv")

    wb = Workbook()
    ws_article = wb.active
    ws_article.title = "Tabla para articulo"

    article_rows = build_base_generation_rows(stats) + build_model_mass_rows(mass_rows)
    add_rows(ws_article, article_rows, title_rows={1, 6}, header_rows={2, 7})
    set_column_widths(ws_article, [13, 7, 5, 7, 8, 8, 8, 8, 8, 8, 9, 10])
    configure_letter_page(ws_article)

    ws_base = wb.create_sheet("Datos base")
    base_rows = [["Estadística descriptiva de agua y boñiga usada por el modelo"]]
    base_rows.append(
        [
            "Variable",
            "Unidad",
            "Promedio",
            "Mediana",
            "Mínimo",
            "Máximo",
            "Desviación estándar",
            "n",
            "Duración del muestreo (días)",
            "Flujo diario",
            "Flujo semanal",
            "Flujo anual",
            "Archivo fuente",
        ]
    )
    for row in stats:
        label = "Agua de lavado" if row["variable"] == "agua" else "Boñiga fresca recolectada"
        base_rows.append(
            [
                label,
                row["unidad"],
                decimal(row["promedio"], 3),
                decimal(row["mediana"], 3),
                decimal(row["minimo"], 3),
                decimal(row["maximo"], 3),
                decimal(row["desviacion_estandar"], 3),
                number(row["n_datos"], 0),
                decimal(row["duracion_muestreo_dias"], 1),
                decimal(row["flujo_por_dia"], 3),
                decimal(row["flujo_por_semana"], 3),
                decimal(row["flujo_por_ano"], 3),
                row["archivo_fuente"],
            ]
        )
    add_rows(ws_base, base_rows, title_rows={1}, header_rows={2})
    set_column_widths(ws_base, [13, 7, 8, 8, 8, 8, 8, 5, 7, 8, 9, 9, 28])
    configure_letter_page(ws_base)

    ws_model = wb.create_sheet("Masa por etapa")
    model_rows = [["Masa anual usada por escenario y etapa"]]
    model_rows.append(
        [
            "Escenario",
            "Etapa",
            "Proceso representado",
            "Fórmula",
            "Boñiga (kg año⁻¹)",
            "Agua (L año⁻¹)",
            "Factor restante A2",
            "Masa total equivalente (kg año⁻¹)",
            "Factor boñiga",
            "Factor agua",
            "Factor masa total",
        ]
    )
    for row in mass_rows:
        escenario = row["escenario"]
        etapa = row["etapa"]
        model_rows.append(
            [
                SCENARIO_LABELS.get(escenario, escenario),
                number(etapa, 0),
                STAGE_LABELS.get((escenario, etapa), ""),
                row["formula"],
                decimal(row["boniga_kg"], 3),
                decimal(row["agua_l"], 3),
                decimal(row["factor_restante_a2"], 6),
                decimal(row["masa_total_kg_eq"], 3),
                decimal(row["factor_boniga_override"], 2),
                decimal(row["factor_agua_override"], 2),
                decimal(row["factor_masa_total_override"], 2),
            ]
        )
    add_rows(ws_model, model_rows, title_rows={1}, header_rows={2})
    set_column_widths(ws_model, [13, 5, 15, 18, 9, 9, 8, 10, 7, 7, 7])
    configure_letter_page(ws_model)

    ws_notes = wb.create_sheet("Notas")
    add_rows(
        ws_notes,
        [
            ["Notas de cálculo"],
            [
                "La generación anual de boñiga fresca corresponde a la fila 'boniga' de "
                "processed/agua_boniga_estadistica_descriptiva.csv."
            ],
            [
                "El flujo diario se obtuvo dividiendo el promedio de cada muestreo entre 3,5 días; "
                "el flujo anual corresponde al flujo diario multiplicado por 365 días."
            ],
            [
                "La masa por escenario y etapa proviene de processed/masa_total_escenario_etapa.csv, "
                "que aplica los factores de asignación definidos en processed/masa_total_factor_overrides.csv."
            ],
            [
                "En el Escenario A, 93 % de la boñiga fresca se asignó a precomposteo/lombricompostaje "
                "y 7 % a la línea de aguas verdes; en el Escenario B se asignó 100 % a purines."
            ],
            [
                "La columna de masa total equivalente mantiene la convención del modelo: 1 L de agua = 1 kg "
                "para sumar agua y boñiga cuando ambas fracciones se manejan juntas."
            ],
        ],
        title_rows={1},
    )
    ws_notes.column_dimensions["A"].width = 90
    configure_letter_page(ws_notes)

    enable_multiline_cells(wb)
    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera una tabla Excel de generación de estiércol usada por el modelo."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Ruta del archivo .xlsx de salida. Valor por defecto: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output if args.output.is_absolute() else ROOT_DIR / args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
