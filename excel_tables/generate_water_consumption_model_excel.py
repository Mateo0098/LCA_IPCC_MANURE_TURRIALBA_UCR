from __future__ import annotations

import argparse
import csv
from copy import copy
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side


ROOT_DIR = Path(__file__).resolve().parents[1]
PROCESSED_DIR = ROOT_DIR / "processed"
DEFAULT_OUTPUT = Path(__file__).resolve().parent / "tabla_consumo_agua_modelo_precision_2_decimales.xlsx"
TABLE_FONT_SIZE = 8
TITLE_FONT_SIZE = 9
MEASUREMENT_DIGITS = 1
MODEL_MASS_DIGITS = 2
WATER_MODEL_DIGITS = 1

STAGE_LABELS = {
    ("A", "1"): "Precomposteo",
    ("A", "2"): "Lombricompostaje",
    ("A", "3"): "Almacenamiento de aguas verdes",
    ("A", "4"): "Aplicación de aguas verdes en campo",
    ("B", "1"): "Almacenamiento de purines",
    ("B", "2"): "Aplicación directa de purines en campo",
}

SCENARIO_LABELS = {
    "A": "A - lombricompostaje y aguas verdes",
    "B": "B - aplicación directa de purines",
}


def read_csv(name: str) -> list[dict[str, str]]:
    path = PROCESSED_DIR / name
    with path.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def number(value: str | float | int, digits: int = 3) -> float | int | str:
    if value in (None, ""):
        return ""
    parsed = round(float(value), digits)
    return int(parsed) if parsed.is_integer() else parsed


def decimal(value: str | float | int, digits: int = 3) -> str:
    if value in (None, ""):
        return ""
    return f"{float(value):.{digits}f}".replace(".", ",")


def add_rows(
    ws,
    rows: list[list[Any]],
    title_rows: set[int] | None = None,
    header_rows: set[int] | None = None,
) -> None:
    title_rows = title_rows or set()
    header_rows = header_rows or set()
    horizontal_line = Side(style="thin", color="000000")
    row_border = Border(bottom=horizontal_line)
    max_columns = max(len(row) for row in rows)

    for row_index, row in enumerate(rows, start=1):
        ws.append(row)
        if not any(value not in (None, "") for value in row):
            continue
        for column_index in range(1, max_columns + 1):
            cell = ws.cell(row=row_index, column=column_index)
            cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
            cell.border = row_border
            cell.font = Font(size=TABLE_FONT_SIZE)
            if row_index in title_rows:
                cell.font = Font(bold=True, size=TITLE_FONT_SIZE)
            if row_index in header_rows:
                cell.font = Font(bold=True, size=TABLE_FONT_SIZE)


def set_column_widths(ws, widths: list[float]) -> None:
    for column_index, width in enumerate(widths, start=1):
        column_letter = ws.cell(row=1, column=column_index).column_letter
        ws.column_dimensions[column_letter].width = width


def configure_letter_page(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = 1
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.5
    ws.page_margins.right = 0.5
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def enable_multiline_cells(wb: Workbook) -> None:
    for ws in wb.worksheets:
        for row in ws.iter_rows(
            min_row=1,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        ):
            for cell in row:
                alignment = copy(cell.alignment)
                alignment.wrap_text = True
                alignment.horizontal = "center"
                alignment.vertical = "center"
                cell.alignment = alignment


def build_base_consumption_rows(stats: list[dict[str, str]]) -> list[list[Any]]:
    water = next(row for row in stats if row["variable"].strip().lower() == "agua")
    return [
        ["Tabla 1. Consumo de agua utilizado como entrada del modelo"],
        [
            "Variable",
            "n",
            "Duración del muestreo (días)",
            "Promedio por muestreo (L)",
            "Mediana (L)",
            "Mínimo (L)",
            "Máximo (L)",
            "Desviación estándar (L)",
            "Flujo diario (L día⁻¹)",
            "Flujo semanal (L semana⁻¹)",
            "Flujo anual usado en el modelo (L año⁻¹)",
        ],
        [
            "Agua de lavado",
            number(water["n_datos"], 0),
            decimal(water["duracion_muestreo_dias"], 1),
            decimal(water["promedio"], MEASUREMENT_DIGITS),
            decimal(water["mediana"], MEASUREMENT_DIGITS),
            decimal(water["minimo"], MEASUREMENT_DIGITS),
            decimal(water["maximo"], MEASUREMENT_DIGITS),
            decimal(water["desviacion_estandar"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_dia"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_semana"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_ano"], WATER_MODEL_DIGITS),
        ],
    ]


def build_model_water_rows(mass_rows: list[dict[str, str]]) -> list[list[Any]]:
    rows = [
        [],
        [],
        ["Tabla 2. Consumo anual de agua asignado a cada escenario y etapa del modelo"],
        [
            "Escenario",
            "Etapa",
            "Proceso representado",
            "Factor de asignación de boñiga",
            "Factor de asignación de agua",
            "Agua incorporada (L año⁻¹)",
            "Boñiga fresca usada por el modelo (kg año⁻¹)",
            "Masa total equivalente (kg año⁻¹)",
            "Origen del valor",
        ],
    ]
    for row in mass_rows:
        escenario = row["escenario"]
        etapa = row["etapa"]
        rows.append(
            [
                SCENARIO_LABELS.get(escenario, escenario),
                number(etapa, 0),
                STAGE_LABELS.get((escenario, etapa), ""),
                decimal(row["factor_boniga_override"], 2),
                decimal(row["factor_agua_override"], 2),
                decimal(row["agua_l"], WATER_MODEL_DIGITS),
                decimal(row["boniga_kg"], MODEL_MASS_DIGITS),
                decimal(row["masa_total_kg_eq"], MODEL_MASS_DIGITS),
                "Factor de asignación del modelo",
            ]
        )
    return rows


def build_workbook() -> Workbook:
    stats = read_csv("agua_boniga_estadistica_descriptiva.csv")
    mass_rows = read_csv("masa_total_escenario_etapa.csv")

    wb = Workbook()
    ws_article = wb.active
    ws_article.title = "Tabla para articulo"

    article_rows = build_base_consumption_rows(stats) + build_model_water_rows(mass_rows)
    add_rows(ws_article, article_rows, title_rows={1, 6}, header_rows={2, 7})
    set_column_widths(ws_article, [13, 5, 8, 9, 8, 8, 8, 8, 9, 10, 11])
    configure_letter_page(ws_article)

    ws_base = wb.create_sheet("Datos base")
    base_rows = [["Estadística descriptiva del consumo de agua usado por el modelo"]]
    base_rows.append(
        [
            "Variable",
            "n",
            "Duración del muestreo (días)",
            "Promedio por muestreo (L)",
            "Mediana (L)",
            "Mínimo (L)",
            "Máximo (L)",
            "Desviación estándar (L)",
            "Flujo diario (L día⁻¹)",
            "Flujo semanal (L semana⁻¹)",
            "Flujo anual (L año⁻¹)",
            "Archivo fuente",
        ]
    )
    water = next(row for row in stats if row["variable"].strip().lower() == "agua")
    base_rows.append(
        [
            "Agua de lavado",
            number(water["n_datos"], 0),
            decimal(water["duracion_muestreo_dias"], 1),
            decimal(water["promedio"], MEASUREMENT_DIGITS),
            decimal(water["mediana"], MEASUREMENT_DIGITS),
            decimal(water["minimo"], MEASUREMENT_DIGITS),
            decimal(water["maximo"], MEASUREMENT_DIGITS),
            decimal(water["desviacion_estandar"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_dia"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_semana"], MEASUREMENT_DIGITS),
            decimal(water["flujo_por_ano"], WATER_MODEL_DIGITS),
            water["archivo_fuente"],
        ]
    )
    add_rows(ws_base, base_rows, title_rows={1}, header_rows={2})
    set_column_widths(ws_base, [13, 5, 8, 9, 8, 8, 8, 8, 9, 10, 11, 28])
    configure_letter_page(ws_base)

    ws_model = wb.create_sheet("Agua por etapa")
    model_rows = [["Consumo anual de agua usado por escenario y etapa"]]
    model_rows.append(
        [
            "Escenario",
            "Etapa",
            "Proceso representado",
            "Fórmula",
            "Agua (L año⁻¹)",
            "Factor agua",
            "Boñiga (kg año⁻¹)",
            "Masa total equivalente (kg año⁻¹)",
        ]
    )
    for row in mass_rows:
        escenario = row["escenario"]
        etapa = row["etapa"]
        model_rows.append(
            [
                SCENARIO_LABELS.get(escenario, escenario),
                number(etapa, 0),
                STAGE_LABELS.get((escenario, etapa), ""),
                row["formula"],
                decimal(row["agua_l"], WATER_MODEL_DIGITS),
                decimal(row["factor_agua_override"], 2),
                decimal(row["boniga_kg"], MODEL_MASS_DIGITS),
                decimal(row["masa_total_kg_eq"], MODEL_MASS_DIGITS),
            ]
        )
    add_rows(ws_model, model_rows, title_rows={1}, header_rows={2})
    set_column_widths(ws_model, [13, 5, 15, 18, 9, 7, 9, 10])
    configure_letter_page(ws_model)

    ws_notes = wb.create_sheet("Notas")
    add_rows(
        ws_notes,
        [
            ["Notas de cálculo"],
            [
                "El consumo anual de agua corresponde a la fila 'agua' de "
                "processed/agua_boniga_estadistica_descriptiva.csv."
            ],
            [
                "El flujo diario se obtuvo dividiendo el promedio de cada muestreo entre 3,5 días; "
                "el flujo anual corresponde al flujo diario multiplicado por 365 días."
            ],
            [
                "La asignación por escenario y etapa proviene de processed/masa_total_escenario_etapa.csv, "
                "que aplica los factores definidos en processed/masa_total_factor_overrides.csv."
            ],
            [
                "En el modelo, el agua se incorporó en las etapas A4 y B2, con factor de asignación de agua igual a 1."
            ],
            [
                "Los valores medidos de agua se reportaron con un decimal, porque las mediciones originales "
                "se registraron en L con un decimal; las masas equivalentes derivadas se mantuvieron con dos decimales."
            ],
        ],
        title_rows={1},
    )
    ws_notes.column_dimensions["A"].width = 90
    configure_letter_page(ws_notes)

    enable_multiline_cells(wb)
    return wb


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Genera una tabla Excel de consumo de agua usado por el modelo."
    )
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"Ruta del archivo .xlsx de salida. Valor por defecto: {DEFAULT_OUTPUT}",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output = args.output if args.output.is_absolute() else ROOT_DIR / args.output
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = build_workbook()
    workbook.save(output)
    print(output)


if __name__ == "__main__":
    main()
