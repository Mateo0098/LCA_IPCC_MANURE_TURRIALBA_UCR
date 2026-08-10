#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import statistics
import re
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

import openpyxl

NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"
Q = f"{{{NS_MAIN}}}"
RQ = f"{{{NS_REL}}}"
PQ = f"{{{NS_PKG}}}"


def ascii_fold(text: str) -> str:
    return unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")


def normalize_key(text: str) -> str:
    text = text.replace("+", " mas ").replace("-", " ")
    text = ascii_fold(text).strip().lower()
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^a-z0-9_]+", "", text)
    return text.strip("_")


def canonical_header(text: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", ascii_fold(text).upper())


def canonical_unit(text: str) -> str:
    cleaned = re.sub(r"\s+", "", ascii_fold(text).lower())
    cleaned = cleaned.replace("\\", "/")
    return cleaned


def unit_suffix(unit: str) -> str:
    c = canonical_unit(unit)
    if not c:
        return ""
    if c == "%":
        return "pct"
    c = c.replace("%", "pct")
    c = re.sub(r"[^a-z0-9]+", "_", c)
    return c.strip("_")


def key_with_unit(base_key: str, unit: str) -> str:
    suffix = unit_suffix(unit)
    return f"{base_key}_{suffix}" if suffix else base_key


def to_number_if_possible(value: str):
    value = value.strip()
    if not value:
        return ""
    if re.fullmatch(r"[+-]?\d+", value):
        try:
            return int(value)
        except ValueError:
            return value
    if re.fullmatch(r"[+-]?(?:\d+\.?\d*|\d*\.?\d+)(?:[eE][+-]?\d+)?", value):
        try:
            return float(value)
        except ValueError:
            return value
    return value


def to_float_or_none(value: object) -> Optional[float]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def infer_treatment(sample_id: str) -> str:
    text = str(sample_id).strip()
    return re.sub(r"\s+\d+$", "", text)


def spanish_long_date_to_ddmmyyyy(text: str) -> str:
    month_map = {
        "enero": "01",
        "febrero": "02",
        "marzo": "03",
        "abril": "04",
        "mayo": "05",
        "junio": "06",
        "julio": "07",
        "agosto": "08",
        "septiembre": "09",
        "setiembre": "09",
        "octubre": "10",
        "noviembre": "11",
        "diciembre": "12",
    }
    m = re.search(r"(\d{1,2})\s+de\s+([a-zA-Záéíóúñ]+)\s+del\s+(\d{4})", text, flags=re.IGNORECASE)
    if not m:
        return ""
    day = f"{int(m.group(1)):02d}"
    month = month_map.get(ascii_fold(m.group(2)).lower(), "")
    year = m.group(3)
    if not month:
        return ""
    return f"{day}/{month}/{year}"


def to_float_from_decimal_comma(value: str) -> Optional[float]:
    text = str(value).strip().replace(".", "").replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def extract_fresh_manure_records_from_pdf(pdf_path: Path) -> List[Dict[str, object]]:
    try:
        from pypdf import PdfReader
    except Exception:
        return []

    try:
        reader = PdfReader(str(pdf_path))
    except Exception:
        return []

    text = "\n".join((page.extract_text() or "") for page in reader.pages)
    if not text.strip():
        return []

    sample_map: Dict[str, str] = {}
    for key, num in re.findall(r"\b([AB])\.\s*Fresco\s+(\d+)\b", text, flags=re.IGNORECASE):
        sample_map[key.upper()] = num

    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    replicas_by_sample: Dict[str, List[float]] = {"A": [], "B": []}
    current_sample = ""
    for line in lines:
        upper = line.upper()
        if upper == "A":
            current_sample = "A"
            continue
        if upper == "B":
            current_sample = "B"
            continue
        if not current_sample:
            continue
        m = re.match(r"^([123])\s+([0-9]+,[0-9]+)\b", line)
        if m:
            value = to_float_from_decimal_comma(m.group(2))
            if value is not None:
                replicas_by_sample[current_sample].append(value)

    # Fallback to medians if replicas could not be parsed.
    if not any(replicas_by_sample.values()):
        median_matches = re.findall(r"Mediana\s+([0-9]+,[0-9]+)", text, flags=re.IGNORECASE)
        if not median_matches:
            return []
        for idx, med in enumerate(median_matches):
            key = "A" if idx == 0 else "B"
            v = to_float_from_decimal_comma(med)
            if v is not None:
                replicas_by_sample[key] = [v]

    report_date = spanish_long_date_to_ddmmyyyy(text)

    records: List[Dict[str, object]] = []
    for key in ("A", "B"):
        values = replicas_by_sample.get(key, [])
        if not values:
            continue
        n_pct = sum(values) / len(values)
        sample_suffix = sample_map.get(key, "2" if key == "A" else "3")
        records.append(
            {
                "date": report_date,
                "sample_id": f"ESTIERCOL FRESCO {sample_suffix}",
                "analysis_type": "N total (Kjeldahl)",
                "n_pctmasa": n_pct,
                "n_total_porcentaje": n_pct,
                "n_total_mg_kg": n_pct * 10000.0,
            }
        )
    return records


def is_n_total_header(header: str) -> bool:
    return canonical_header(header) in {"NTOTAL", "NITROGENOTOTAL"}


def col_to_index(ref: str) -> int:
    letters = "".join(ch for ch in ref if ch.isalpha())
    result = 0
    for ch in letters:
        result = result * 26 + (ord(ch.upper()) - 64)
    return result


def parse_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    shared: List[str] = []
    for item in root.findall(f"{Q}si"):
        shared.append("".join((node.text or "") for node in item.findall(f".//{Q}t")))
    return shared


def read_cell_value(cell: ET.Element, shared: List[str]) -> Optional[str]:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        node = cell.find(f"{Q}is/{Q}t")
        return "" if node is None else (node.text or "")
    node = cell.find(f"{Q}v")
    if node is None:
        return None
    raw = node.text or ""
    if cell_type == "s" and raw.isdigit():
        idx = int(raw)
        if 0 <= idx < len(shared):
            return shared[idx]
    return raw


def parse_workbook(zf: zipfile.ZipFile) -> List[Tuple[str, str]]:
    workbook = ET.fromstring(zf.read("xl/workbook.xml"))
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rel_map = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall(f"{PQ}Relationship")
    }
    sheets: List[Tuple[str, str]] = []
    for sheet in workbook.findall(f".//{Q}sheets/{Q}sheet"):
        rid = sheet.attrib.get(f"{RQ}id", "")
        target = rel_map.get(rid)
        name = sheet.attrib.get("name", "")
        if target:
            sheets.append((name, f"xl/{target}"))
    return sheets


def read_sheet(
    xml_bytes: bytes, shared: List[str]
) -> Tuple[Dict[int, Dict[int, str]], Dict[int, Set[int]]]:
    root = ET.fromstring(xml_bytes)
    rows: Dict[int, Dict[int, str]] = {}
    formula_cells: Dict[int, Set[int]] = {}

    for row in root.findall(f".//{Q}sheetData/{Q}row"):
        row_num = int(row.attrib.get("r", "0"))
        row_data: Dict[int, str] = {}
        row_formulas: Set[int] = set()
        for cell in row.findall(f"{Q}c"):
            ref = cell.attrib.get("r", "")
            if not ref:
                continue
            col = col_to_index(ref)
            value = read_cell_value(cell, shared)
            if value is not None:
                row_data[col] = value
            if cell.find(f"{Q}f") is not None:
                row_formulas.add(col)
        if row_data:
            rows[row_num] = row_data
        if row_formulas:
            formula_cells[row_num] = row_formulas
    return rows, formula_cells


def get_label_value(rows: Dict[int, Dict[int, str]], label: str) -> str:
    target = ascii_fold(label).strip().upper()
    for row_num in sorted(rows):
        row = rows[row_num]
        for col in sorted(row):
            text = ascii_fold(str(row[col] or "")).strip().upper()
            if text == target:
                for next_col in sorted(c for c in row if c > col):
                    val = str(row[next_col] or "").strip()
                    if val:
                        return val
                next_row = rows.get(row_num + 1, {})
                val = str(next_row.get(col, "") or "").strip()
                if val:
                    return val
    return ""


def find_table_header_row(rows: Dict[int, Dict[int, str]]) -> Optional[int]:
    for row_num in sorted(rows):
        values = {ascii_fold(str(v)).strip().upper() for v in rows[row_num].values() if str(v).strip()}
        if "ID USUARIO" in values:
            return row_num
    return None


def find_best_sheet(
    sheets_data: Dict[str, Tuple[Dict[int, Dict[int, str]], Dict[int, Set[int]]]]
) -> Optional[str]:
    for sheet_name, (rows, _) in sheets_data.items():
        if find_table_header_row(rows) is not None:
            return sheet_name
    return None


def detect_density(values_by_key: Dict[str, object], units_by_key: Dict[str, str]) -> Optional[float]:
    dens = values_by_key.get("densidad")
    if not isinstance(dens, (int, float)):
        return None
    unit = canonical_unit(units_by_key.get("densidad", ""))
    if unit in {"g/ml", "kg/l", ""}:
        return float(dens)
    return None


def compute_n_total_columns(
    values_by_key: Dict[str, object],
    units_by_key: Dict[str, str],
) -> Tuple[object, object]:
    density = detect_density(values_by_key, units_by_key)

    components: List[Tuple[float, str]] = []
    for key in ("n_nh4", "n_no3", "n_ureico"):
        value = values_by_key.get(key)
        if isinstance(value, (int, float)):
            components.append((float(value), canonical_unit(units_by_key.get(key, ""))))

    if not components:
        value = values_by_key.get("n")
        if isinstance(value, (int, float)):
            components.append((float(value), canonical_unit(units_by_key.get("n", ""))))

    if not components:
        return "", ""

    total_value = float(sum(v for v, _ in components))
    unit_set = {u for _, u in components if u}

    if "%" in unit_set or any("masa" in u for u in unit_set):
        n_total_porcentaje = total_value
        n_total_mg_kg = total_value * 10000.0
        return n_total_porcentaje, n_total_mg_kg

    if "mg/kg" in unit_set:
        n_total_mg_kg = total_value
        n_total_porcentaje = total_value / 10000.0
        return n_total_porcentaje, n_total_mg_kg

    if "mg/l" in unit_set:
        if density is None or density == 0:
            return "", ""
        n_total_mg_kg = total_value / density
        n_total_porcentaje = n_total_mg_kg / 10000.0
        return n_total_porcentaje, n_total_mg_kg

    return "", ""


def extract_records_from_sheet(
    rows: Dict[int, Dict[int, str]],
    formula_cells: Dict[int, Set[int]],
) -> List[Dict[str, object]]:
    header_row_num = find_table_header_row(rows)
    if header_row_num is None:
        return []

    header_row = rows[header_row_num]
    header_cols = sorted(header_row.keys())
    headers = {col: str(header_row[col]).strip() for col in header_cols}
    canonical_by_col = {col: canonical_header(headers[col]) for col in header_cols}
    unit_row = rows.get(header_row_num - 1, {})

    raw_units_by_col: Dict[int, str] = {}
    for col in header_cols:
        unit_text = str(unit_row.get(col, "")).strip()
        if unit_text:
            raw_units_by_col[col] = unit_text
    # forward fill units for merged/compact unit rows (common in these lab reports)
    last_unit = ""
    for col in header_cols:
        if col in raw_units_by_col:
            last_unit = raw_units_by_col[col]
        elif last_unit:
            raw_units_by_col[col] = last_unit

    measurement_cols: List[int] = []
    for col in header_cols:
        canonical = canonical_by_col[col]
        if canonical in {"", "IDUSUARIO", "IDLAB"}:
            continue
        if is_n_total_header(headers[col]):
            continue
        has_formula = any(col in formula_cells.get(rn, set()) for rn in rows if rn > header_row_num)
        if has_formula:
            continue
        measurement_cols.append(col)

    id_usuario_col = next((col for col in header_cols if canonical_by_col[col] == "IDUSUARIO"), None)
    if id_usuario_col is None:
        return []

    fecha = get_label_value(rows, "EMISION DE REPORTE:") or get_label_value(rows, "FECHA RECEPCION:")
    tipo_analisis = get_label_value(rows, "ANALISIS:")

    records: List[Dict[str, object]] = []
    for row_num in sorted(rn for rn in rows if rn > header_row_num):
        row = rows[row_num]
        id_usuario = str(row.get(id_usuario_col, "")).strip()
        id_ascii = ascii_fold(id_usuario).upper()
        if not id_usuario:
            continue
        if "ULTIMA LINEA" in id_ascii:
            break

        record: Dict[str, object] = {
            "date": fecha.strip(),
            "sample_id": id_usuario,
            "analysis_type": tipo_analisis.strip(),
        }
        values_by_key: Dict[str, object] = {}
        units_by_key: Dict[str, str] = {}

        for col in measurement_cols:
            base_key = normalize_key(headers[col])
            if not base_key:
                continue
            value = str(row.get(col, "")).strip()
            parsed = to_number_if_possible(value) if value else ""
            unit = raw_units_by_col.get(col, "")
            output_key = key_with_unit(base_key, unit)
            record[output_key] = parsed
            values_by_key[base_key] = parsed
            if unit:
                units_by_key[base_key] = unit

        n_total_porcentaje, n_total_mg_kg = compute_n_total_columns(values_by_key, units_by_key)
        record["n_total_porcentaje"] = n_total_porcentaje
        record["n_total_mg_kg"] = n_total_mg_kg
        records.append(record)
    return records


def extract_from_workbook(path: Path) -> List[Dict[str, object]]:
    with zipfile.ZipFile(path) as zf:
        shared = parse_shared_strings(zf)
        sheets = parse_workbook(zf)
        sheets_data: Dict[str, Tuple[Dict[int, Dict[int, str]], Dict[int, Set[int]]]] = {}
        for name, target in sheets:
            sheets_data[name] = read_sheet(zf.read(target), shared)

    selected = find_best_sheet(sheets_data)
    if not selected:
        return []
    rows, formulas = sheets_data[selected]
    return extract_records_from_sheet(rows, formulas)


def discover_workbooks(input_path: Path) -> List[Path]:
    if input_path.is_file():
        return [input_path]
    return sorted(
        p for p in input_path.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")
    )


def write_output_csv(base_path: Path, prefix: str, records: List[Dict[str, object]]) -> Path:
    csv_path = base_path / f"{prefix}.csv"

    fieldnames = sorted({k for record in records for k in record.keys()})
    with csv_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for record in records:
            writer.writerow(record)

    return csv_path


def build_treatment_summary(records: List[Dict[str, object]]) -> List[Dict[str, object]]:
    groups: Dict[str, Dict[str, object]] = {}
    for row in records:
        treatment = infer_treatment(str(row.get("sample_id", "")))
        if treatment not in groups:
            groups[treatment] = {
                "n_total_mg_kg": [],
                "n_total_porcentaje": [],
                "dates": set(),
            }

        n_total = to_float_or_none(row.get("n_total_mg_kg", ""))
        n_pct = to_float_or_none(row.get("n_total_porcentaje", ""))
        date_value = str(row.get("date", "")).strip()
        if n_total is not None:
            groups[treatment]["n_total_mg_kg"].append(n_total)  # type: ignore[index]
        if n_pct is not None:
            groups[treatment]["n_total_porcentaje"].append(n_pct)  # type: ignore[index]
        if date_value:
            groups[treatment]["dates"].add(date_value)  # type: ignore[index]

    summary_rows: List[Dict[str, object]] = []
    for treatment in sorted(groups):
        totals = groups[treatment]["n_total_mg_kg"]  # type: ignore[index]
        pcts = groups[treatment]["n_total_porcentaje"]  # type: ignore[index]
        dates = sorted(groups[treatment]["dates"])  # type: ignore[index]
        n_samples = max(len(totals), len(pcts))
        mean_total = sum(totals) / len(totals) if totals else ""
        mean_pct = sum(pcts) / len(pcts) if pcts else ""
        median_total = statistics.median(totals) if totals else ""
        min_total = min(totals) if totals else ""
        max_total = max(totals) if totals else ""
        median_pct = statistics.median(pcts) if pcts else ""
        summary_rows.append(
            {
                "treatment": treatment,
                "date": " | ".join(dates),
                "n_samples": n_samples,
                "mean_n_total_mg_kg": mean_total,
                "mean_n_percentage": mean_pct,
                "n_total_mg_kg_median": median_total,
                "n_total_mg_kg_min": min_total,
                "n_total_mg_kg_max": max_total,
                "n_total_pct_median": median_pct,
            }
        )
    return summary_rows


def write_treatment_summary_csv(base_path: Path, prefix: str, summary_rows: List[Dict[str, object]]) -> Path:
    summary_path = base_path / f"{prefix}_treatment_summary.csv"
    with summary_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=[
                "treatment",
                "date",
                "n_samples",
                "mean_n_total_mg_kg",
                "mean_n_percentage",
                "n_total_mg_kg_median",
                "n_total_mg_kg_min",
                "n_total_mg_kg_max",
                "n_total_pct_median",
            ],
        )
        writer.writeheader()
        for row in summary_rows:
            writer.writerow(row)
    return summary_path


# ---------------------------------------------------------------------------
# Ingestión normalizada multijornada. La interfaz histórica anterior se
# conserva para no alterar todavía ningún consumidor del modelo ACV.
# ---------------------------------------------------------------------------

NORMALIZED_COMMON_FIELDS = [
    "jornada_muestreo", "fecha_muestreo", "fecha_recepcion", "fecha_analisis",
    "tipo_material", "identificador_muestra", "identificador_muestra_origen",
    "repeticion_muestra", "replica_analitica", "nivel_observacion", "variable",
    "valor", "unidad", "base_medicion", "incertidumbre", "laboratorio",
    "metodo_analitico", "fuente_metodo_analitico", "archivo_origen",
    "hoja_origen", "celda_o_fila_origen", "id_reporte_laboratorio",
    "uso_modelo", "motivo_uso_modelo", "bandera_calidad",
]


def _fold(value: object) -> str:
    return re.sub(r"\s+", " ", ascii_fold(str(value or ""))).strip().upper()


def _to_float_locale(value: object) -> Optional[float]:
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value or "").strip().replace(" ", "")
    if not text or text.startswith("="):
        return None
    if "," in text and "." not in text:
        text = text.replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def _material_code(material: str) -> str:
    return {
        "estiércol fresco": "EF",
        "estiércol precompostado": "EP",
        "aguas verdes": "AV",
        "purines": "PU",
    }[material]


def _sample_number(origin_id: str, jornada: str, fallback: int) -> int:
    folded = _fold(origin_id)
    if jornada != "M1":
        match = re.search(r"(?:^|\D)2\s*[-,]\s*(\d+)\s*$", folded)
        if match:
            return int(match.group(1))
    match = re.search(r"(\d+)\s*$", folded)
    return int(match.group(1)) if match else fallback


def _find_result_sheet(workbook: openpyxl.Workbook):
    candidates = []
    for worksheet in workbook.worksheets:
        header_row = None
        for row in worksheet.iter_rows():
            if any(_fold(cell.value) == "ID USUARIO" for cell in row):
                header_row = row[0].row
                break
        if header_row is None:
            continue
        result_rows = 0
        for row_num in range(header_row + 1, min(worksheet.max_row, header_row + 30) + 1):
            value = worksheet.cell(row_num, 1).value
            if value and "ULTIMA LINEA" in _fold(value):
                break
            if value and ("LIQ:" in _fold(value) or "SOL:" in _fold(value)):
                result_rows += 1
        if result_rows:
            candidates.append((result_rows, worksheet, header_row))
    if not candidates:
        raise ValueError("No se encontró una tabla CIA con filas reales de resultados")
    candidates.sort(key=lambda item: item[0], reverse=True)
    return candidates[0][1], candidates[0][2]


def _label_value(worksheet, labels: set[str]) -> str:
    targets = {_fold(label).rstrip(":") for label in labels}
    for row in worksheet.iter_rows():
        for cell in row:
            if _fold(cell.value).rstrip(":") in targets:
                for col in range(cell.column + 1, worksheet.max_column + 1):
                    value = worksheet.cell(cell.row, col).value
                    if value not in (None, ""):
                        return str(value).strip()
    return ""


def _cia_variable(header: str, material: str) -> str:
    key = canonical_header(header)
    mapping = {
        "N": "N total",
        "NNH4": "N amoniacal",
        "NNO3": "N nítrico",
        "NUREICO": "N ureico",
        "C": "carbono",
        "CN": "relación C/N",
        "DENSIDAD": "densidad",
    }
    if key == "N" and material in {"aguas verdes", "purines"}:
        return "N total"
    return mapping.get(key, header.strip())


def _base_for(variable: str, unit: str) -> str:
    folded_unit = _fold(unit)
    if variable in {"cenizas", "sólidos volátiles"}:
        return "masa seca"
    if "%" in unit or "MASA" in folded_unit:
        return "masa fresca" if variable in {"humedad", "materia seca", "N total", "carbono"} else "masa"
    if "/L" in folded_unit:
        return "volumen"
    return "no especificada"


def _cia_base_for(material: str, variable: str, unit: str) -> str:
    # Los reportes CIA 97600 y 100751 expresan N/C como porcentaje, pero no
    # declaran si el resultado final está referido a base seca o fresca. El
    # secado a 80 °C es preparación de muestra y no basta para inferir la base.
    if material == "estiércol precompostado" and variable in {"N total", "carbono"}:
        return "no especificada en el reporte"
    return _base_for(variable, unit)


def _variable_usage(source: Dict[str, object], variable: str) -> Tuple[str, str]:
    if variable in {"densidad", "carbono", "relación C/N"}:
        return (
            "solo_caracterizacion",
            "Variable conservada para caracterización y trazabilidad; no es un parámetro consumido actualmente por el modelo ACV.",
        )
    return str(source["uso_modelo"]), str(source["motivo_uso"])


def extract_cia_normalized(source: Dict[str, object], project_root: Path) -> List[Dict[str, object]]:
    path = Path(source["absolute_path"])
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    worksheet, header_row = _find_result_sheet(workbook)
    headers = {cell.column: str(cell.value).strip() for cell in worksheet[header_row] if cell.value not in (None, "")}
    id_user_col = next(col for col, value in headers.items() if _fold(value) == "ID USUARIO")
    id_lab_col = next((col for col, value in headers.items() if _fold(value).replace(" ", "") == "IDLAB"), None)

    units: Dict[int, str] = {}
    last_unit = ""
    for col in sorted(headers):
        raw = worksheet.cell(header_row - 1, col).value
        if raw not in (None, ""):
            last_unit = str(raw).strip()
        if last_unit:
            units[col] = last_unit

    report_id = _label_value(worksheet, {"Nº DE REPORTE", "N° DE REPORTE"})
    reception = _label_value(worksheet, {"FECHA RECEPCIÓN", "FECHA DE RECEPCIÓN"})
    emission = _label_value(worksheet, {"EMISIÓN DE REPORTE"})
    records: List[Dict[str, object]] = []
    sample_index = 0
    for row_num in range(header_row + 1, worksheet.max_row + 1):
        origin_id = str(worksheet.cell(row_num, id_user_col).value or "").strip()
        if "ULTIMA LINEA" in _fold(origin_id):
            break
        if not origin_id:
            continue
        sample_index += 1
        sample_number = _sample_number(origin_id, str(source["jornada"]), sample_index)
        normalized_id = f'{source["jornada"]}-{_material_code(str(source["material"]))}-{sample_number}'
        id_lab = str(worksheet.cell(row_num, id_lab_col).value or "").strip() if id_lab_col else ""
        for col, header in headers.items():
            if col in {id_user_col, id_lab_col}:
                continue
            value = _to_float_locale(worksheet.cell(row_num, col).value)
            if value is None:
                continue
            variable = _cia_variable(header, str(source["material"]))
            # Las columnas derivadas o de resumen nunca son observaciones primarias.
            if variable not in {"N total", "N amoniacal", "N nítrico", "N ureico", "carbono", "relación C/N", "densidad"}:
                continue
            unit = units.get(col, "")
            method = "densidad reportada" if variable == "densidad" else source["metodo"]
            method_source = (
                "Valor de densidad consignado en el informe CIA; el procedimiento no se describe."
                if variable == "densidad" else source["fuente_metodo"]
            )
            variable_usage, variable_usage_reason = _variable_usage(source, variable)
            records.append({
                "jornada_muestreo": source["jornada"], "fecha_muestreo": "",
                "fecha_recepcion": reception, "fecha_analisis": emission,
                "tipo_material": source["material"], "identificador_muestra": normalized_id,
                "identificador_muestra_origen": origin_id, "repeticion_muestra": sample_number,
                "replica_analitica": "", "nivel_observacion": "muestra_compuesta",
                "variable": variable, "valor": value, "unidad": unit,
                "base_medicion": _cia_base_for(str(source["material"]), variable, unit), "incertidumbre": "",
                "laboratorio": source["laboratorio"], "metodo_analitico": method,
                "fuente_metodo_analitico": method_source,
                "archivo_origen": str(path.relative_to(project_root)).replace("\\", "/"),
                "hoja_origen": worksheet.title, "celda_o_fila_origen": f"fila {row_num}",
                "id_reporte_laboratorio": report_id or id_lab,
                "uso_modelo": variable_usage, "motivo_uso_modelo": variable_usage_reason,
                "bandera_calidad": "",
            })
    return records


def extract_lasa_normalized(source: Dict[str, object], project_root: Path) -> List[Dict[str, object]]:
    from pypdf import PdfReader

    path = Path(source["absolute_path"])
    reader = PdfReader(str(path))
    pages = [page.extract_text() or "" for page in reader.pages]
    text = "\n".join(pages)
    descriptions = {
        key.upper(): value.rstrip(".")
        for key, value in re.findall(r"\b([ABC])\.\s*(Fresco\s+\d+(?:,\d+)?)", text, flags=re.IGNORECASE)
    }
    sample_date_match = re.search(r"Muestreo\s+(\d{1,2}/\d{1,2}/\d{4})", text, flags=re.IGNORECASE)
    reception_match = re.search(r"Fecha de recepción:\s*(\d{1,2}/\d{1,2}/\d{2,4})", text, flags=re.IGNORECASE)
    analysis_match = re.search(r"Fecha de análisis:\s*(.+?\d{4})", text, flags=re.IGNORECASE)
    report_match = re.search(r"Informe de Analisis\s+N\S*\s*([0-9]+-[0-9]+)", ascii_fold(text), flags=re.IGNORECASE)
    results_text = text.split("Cuadro I.", 1)[-1].split("Descripción del procedimiento", 1)[0]
    starts = list(re.finditer(r"(?:^|\s)([ABC])\s+(?=(?:Contenido|1\s+\d))", results_text, flags=re.IGNORECASE))
    records: List[Dict[str, object]] = []
    for index, start in enumerate(starts):
        key = start.group(1).upper()
        end = starts[index + 1].start() if index + 1 < len(starts) else len(results_text)
        block = results_text[start.end():end]
        sample_number = index + 1
        origin = f"{key}. {descriptions.get(key, 'Fresco sin descripción')}"
        normalized_id = f'{source["jornada"]}-EF-{sample_number}'
        for match in re.finditer(r"\b([123])\s+([0-9]+,[0-9]+)\s*±\s*([0-9]+,[0-9]+)", block):
            records.append({
                "jornada_muestreo": source["jornada"],
                "fecha_muestreo": sample_date_match.group(1) if sample_date_match else "",
                "fecha_recepcion": reception_match.group(1) if reception_match else "",
                "fecha_analisis": analysis_match.group(1).strip() if analysis_match else "",
                "tipo_material": source["material"], "identificador_muestra": normalized_id,
                "identificador_muestra_origen": origin, "repeticion_muestra": sample_number,
                "replica_analitica": int(match.group(1)), "nivel_observacion": "replica_analitica",
                "variable": "N total", "valor": to_float_from_decimal_comma(match.group(2)),
                "unidad": "% masa", "base_medicion": "masa fresca",
                "incertidumbre": to_float_from_decimal_comma(match.group(3)),
                "laboratorio": source["laboratorio"], "metodo_analitico": source["metodo"],
                "fuente_metodo_analitico": source["fuente_metodo"],
                "archivo_origen": str(path.relative_to(project_root)).replace("\\", "/"),
                "hoja_origen": "página 1", "celda_o_fila_origen": f"muestra {key}, réplica {match.group(1)}",
                "id_reporte_laboratorio": report_match.group(1) if report_match else "",
                "uso_modelo": source["uso_modelo"], "motivo_uso_modelo": source["motivo_uso"],
                "bandera_calidad": "",
            })
    return records


def main() -> int:
    project_root = Path(__file__).resolve().parent.parent
    preferred_default = project_root / "Academic_documents" / "resultados CIA y LASA muestreo 1"
    default_input = preferred_default if preferred_default.exists() else (project_root / "Academic_documents")

    parser = argparse.ArgumentParser(
        description=(
            "Extract lab results and compute n_total in Python "
            "(without using the Excel n_total column)."
        )
    )
    parser.add_argument(
        "input_path",
        nargs="?",
        default=None,
        help="Input .xlsx file or folder with .xlsx files (default: Academic_documents)",
    )
    parser.add_argument(
        "--out-prefix",
        default="CIA_samples_table",
        help="Output CSV prefix (default: CIA_samples_table)",
    )
    args = parser.parse_args()

    input_path = Path(args.input_path) if args.input_path else default_input
    if not input_path.exists():
        raise SystemExit(f"Input path does not exist: {input_path}")

    workbooks = discover_workbooks(input_path)
    if not workbooks:
        raise SystemExit(f"No .xlsx files found in: {input_path}")

    all_records: List[Dict[str, object]] = []
    for wb in workbooks:
        all_records.extend(extract_from_workbook(wb))

    pdf_dir = input_path if input_path.is_dir() else input_path.parent
    pdf_candidates = sorted(pdf_dir.glob("*Contenido de nitrogeno*.pdf"))
    if pdf_candidates:
        all_records.extend(extract_fresh_manure_records_from_pdf(pdf_candidates[0]))

    if not all_records:
        raise SystemExit("No result rows found to extract.")

    output_dir = project_root / "processed"
    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = write_output_csv(output_dir, args.out_prefix, all_records)
    summary_rows = build_treatment_summary(all_records)
    summary_path = write_treatment_summary_csv(output_dir, args.out_prefix, summary_rows)

    print(f"Rows extracted: {len(all_records)}")
    print(f"CSV: {csv_path}")
    print(f"Summary CSV: {summary_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
