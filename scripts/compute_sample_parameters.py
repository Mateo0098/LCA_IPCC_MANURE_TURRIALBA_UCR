from __future__ import annotations

import csv
import re
import statistics
import xml.etree.ElementTree as ET
import zipfile
import unicodedata
from collections import defaultdict
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import openpyxl

NS = {"a": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
DRYING_ID_RE = re.compile(r"^[AB]\d{2}$")
INCINERATION_ID_RE = re.compile(r"^[AB]I\d{2}$")


def col_to_idx(col_letters: str) -> int:
    idx = 0
    for ch in col_letters:
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def parse_cell_ref(cell_ref: str) -> Tuple[int, int]:
    i = 0
    while i < len(cell_ref) and cell_ref[i].isalpha():
        i += 1
    return col_to_idx(cell_ref[:i]), int(cell_ref[i:])


def load_shared_strings(zf: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in zf.namelist():
        return []
    root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
    strings = []
    for si in root.findall("a:si", NS):
        strings.append("".join(t.text or "" for t in si.findall(".//a:t", NS)))
    return strings


def read_sheet_rows_as_table(
    workbook_path: Path, sheet_xml_path: str
) -> Dict[Tuple[int, int], str]:
    with zipfile.ZipFile(workbook_path) as zf:
        shared_strings = load_shared_strings(zf)
        sheet_root = ET.fromstring(zf.read(sheet_xml_path))

    table: Dict[Tuple[int, int], str] = {}
    for row in sheet_root.findall("a:sheetData/a:row", NS):
        row_num = int(row.attrib["r"])
        for cell in row.findall("a:c", NS):
            col_num, _ = parse_cell_ref(cell.attrib["r"])
            cell_type = cell.attrib.get("t")
            value_node = cell.find("a:v", NS)
            if value_node is None:
                value = ""
            elif cell_type == "s":
                value = shared_strings[int(value_node.text)]
            else:
                value = value_node.text or ""
            table[(row_num, col_num)] = value
    return table


def to_float(text: str) -> Optional[float]:
    if text is None:
        return None
    clean = text.strip().replace(",", ".")
    if not clean:
        return None
    try:
        return float(clean)
    except ValueError:
        return None


def sample_type_from_id(sample_id: str) -> str:
    if sample_id.startswith("A"):
        return "Fresh manure"
    if sample_id.startswith("B"):
        return "Precomposted manure"
    return "Undefined"


def safe_pct(numerator: float, denominator: float) -> Optional[float]:
    if denominator == 0:
        return None
    return (numerator / denominator) * 100.0


def extract_drying_data(table: Dict[Tuple[int, int], str]) -> Dict[str, Dict[str, float]]:
    drying: Dict[str, Dict[str, float]] = {}
    for row in range(2, 58):
        sample_id = table.get((row, 1), "").strip()
        if not DRYING_ID_RE.fullmatch(sample_id):
            continue

        crisol_solo = to_float(table.get((row, 2), ""))
        crisol_con_humeda = to_float(table.get((row, 3), ""))
        crisol_con_seca = to_float(table.get((row, 5), ""))
        if crisol_solo is None or crisol_con_humeda is None or crisol_con_seca is None:
            continue

        peso_muestra_humeda = crisol_con_humeda - crisol_solo
        peso_muestra_seca = crisol_con_seca - crisol_solo

        humedad_pct = safe_pct(peso_muestra_humeda - peso_muestra_seca, peso_muestra_humeda)
        masa_seca_pct = safe_pct(peso_muestra_seca, peso_muestra_humeda)
        if humedad_pct is None or masa_seca_pct is None:
            continue

        drying[sample_id] = {
            "wet_sample_mass_g": peso_muestra_humeda,
            "dry_sample_mass_g": peso_muestra_seca,
            "moisture_content_pct": humedad_pct,
            "dry_matter_pct": masa_seca_pct,
        }
    return drying


def extract_incineration_data(
    table: Dict[Tuple[int, int], str]
) -> Dict[str, Dict[str, float]]:
    incineration: Dict[str, Dict[str, float]] = {}
    for row in range(2, 58):
        sample_id = table.get((row, 1), "").strip()
        if not INCINERATION_ID_RE.fullmatch(sample_id):
            continue

        sample_id_base = sample_id.replace("I", "", 1)
        crisol_solo = to_float(table.get((row, 2), ""))
        crisol_con_seca = to_float(table.get((row, 3), ""))
        crisol_con_ceniza = to_float(table.get((row, 5), ""))
        if crisol_solo is None or crisol_con_seca is None or crisol_con_ceniza is None:
            continue

        peso_muestra_seca_inc_g = crisol_con_seca - crisol_solo
        peso_ceniza_g = crisol_con_ceniza - crisol_solo

        cenizas_pct = safe_pct(peso_ceniza_g, peso_muestra_seca_inc_g)
        if cenizas_pct is None:
            continue
        solidos_volatiles_pct = 100.0 - cenizas_pct

        incineration[sample_id_base] = {
            "dry_sample_mass_for_ash_test_g": peso_muestra_seca_inc_g,
            "ash_mass_g": peso_ceniza_g,
            "ash_content_pct": cenizas_pct,
            "volatile_solids_pct": solidos_volatiles_pct,
        }
    return incineration


def merge_results(
    drying: Dict[str, Dict[str, float]], incineration: Dict[str, Dict[str, float]]
) -> List[Dict[str, object]]:
    merged_rows: List[Dict[str, object]] = []
    all_ids = sorted(set(drying.keys()) | set(incineration.keys()))
    for sample_id in all_ids:
        row: Dict[str, object] = {
            "sampling_date": "2025-11-10",
            "sample_id": sample_id,
            "sample_type": sample_type_from_id(sample_id),
        }
        row.update(drying.get(sample_id, {}))
        row.update(incineration.get(sample_id, {}))
        merged_rows.append(row)
    return merged_rows


def write_csv(rows: Iterable[Dict[str, object]], output_path: Path) -> None:
    columns = [
        "sampling_date",
        "sample_id",
        "sample_type",
        "wet_sample_mass_g",
        "dry_sample_mass_g",
        "moisture_content_pct",
        "dry_matter_pct",
        "dry_sample_mass_for_ash_test_g",
        "ash_mass_g",
        "ash_content_pct",
        "volatile_solids_pct",
    ]

    rows_list = list(rows)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows_list:
            formatted = {}
            for col in columns:
                value = row.get(col)
                if isinstance(value, float):
                    formatted[col] = f"{value:.6f}"
                else:
                    formatted[col] = value if value is not None else ""
            writer.writerow(formatted)


def base_sample_id(sample_id: str) -> str:
    if len(sample_id) >= 2:
        return sample_id[:2]
    return sample_id


def mean_sd_cv(values: List[float]) -> Tuple[Optional[float], Optional[float], Optional[float]]:
    if not values:
        return None, None, None
    mean_value = statistics.mean(values)
    sd_value = statistics.stdev(values) if len(values) > 1 else 0.0
    cv_value = (sd_value / mean_value) * 100.0 if mean_value != 0 else None
    return mean_value, sd_value, cv_value


def summarize_representative_values(rows: List[Dict[str, object]]) -> List[Dict[str, object]]:
    metrics = [
        "moisture_content_pct",
        "dry_matter_pct",
        "ash_content_pct",
        "volatile_solids_pct",
    ]

    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in rows:
        sample_id = str(row.get("sample_id", ""))
        grouped[base_sample_id(sample_id)].append(row)

    summary_rows: List[Dict[str, object]] = []
    for sample_group in sorted(grouped.keys()):
        group_rows = grouped[sample_group]
        sample_type = str(group_rows[0].get("sample_type", "Undefined"))
        summary: Dict[str, object] = {
            "sampling_date": "2025-11-10",
            "sample_group": sample_group,
            "sample_type": sample_type,
            "replicate_count": len(group_rows),
        }
        for metric in metrics:
            values: List[float] = []
            for row in group_rows:
                value = row.get(metric)
                if isinstance(value, (int, float)):
                    values.append(float(value))
            mean_value, sd_value, cv_value = mean_sd_cv(values)
            metric_name = metric.replace("_pct", "")
            summary[f"{metric_name}_mean_pct"] = mean_value
            summary[f"{metric_name}_sd_pct"] = sd_value
            summary[f"{metric_name}_cv_pct"] = cv_value
        summary_rows.append(summary)
    return summary_rows


def write_representative_csv(rows: Iterable[Dict[str, object]], output_path: Path) -> None:
    columns = [
        "sampling_date",
        "sample_group",
        "sample_type",
        "replicate_count",
        "moisture_content_mean_pct",
        "moisture_content_sd_pct",
        "moisture_content_cv_pct",
        "dry_matter_mean_pct",
        "dry_matter_sd_pct",
        "dry_matter_cv_pct",
        "ash_content_mean_pct",
        "ash_content_sd_pct",
        "ash_content_cv_pct",
        "volatile_solids_mean_pct",
        "volatile_solids_sd_pct",
        "volatile_solids_cv_pct",
    ]

    rows_list = list(rows)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows_list:
            formatted = {}
            for col in columns:
                value = row.get(col)
                if isinstance(value, float):
                    formatted[col] = f"{value:.6f}"
                else:
                    formatted[col] = value if value is not None else ""
            writer.writerow(formatted)


def summarize_treatment_values(
    representative_rows: List[Dict[str, object]]
) -> List[Dict[str, object]]:
    grouped: Dict[str, List[Dict[str, object]]] = defaultdict(list)
    for row in representative_rows:
        sample_group = str(row.get("sample_group", ""))
        treatment = sample_group[:1] if sample_group else ""
        if treatment:
            grouped[treatment].append(row)

    metric_bases = [
        "moisture_content",
        "dry_matter",
        "ash_content",
        "volatile_solids",
    ]

    treatment_rows: List[Dict[str, object]] = []
    for treatment in sorted(grouped.keys()):
        rows = grouped[treatment]
        treatment_row: Dict[str, object] = {
            "sampling_date": "2025-11-10",
            "treatment": treatment,
            "sample_type": str(rows[0].get("sample_type", "Undefined")),
            "sample_count": len(rows),
        }
        for metric in metric_bases:
            values: List[float] = []
            for row in rows:
                value = row.get(f"{metric}_mean_pct")
                if isinstance(value, (int, float)):
                    values.append(float(value))
            mean_value, sd_value, cv_value = mean_sd_cv(values)
            treatment_row[f"{metric}_treatment_mean_pct"] = mean_value
            treatment_row[f"{metric}_treatment_sd_pct"] = sd_value
            treatment_row[f"{metric}_treatment_cv_pct"] = cv_value
        treatment_rows.append(treatment_row)
    return treatment_rows


def write_treatment_csv(rows: Iterable[Dict[str, object]], output_path: Path) -> None:
    columns = [
        "sampling_date",
        "treatment",
        "sample_type",
        "sample_count",
        "moisture_content_treatment_mean_pct",
        "moisture_content_treatment_sd_pct",
        "moisture_content_treatment_cv_pct",
        "dry_matter_treatment_mean_pct",
        "dry_matter_treatment_sd_pct",
        "dry_matter_treatment_cv_pct",
        "ash_content_treatment_mean_pct",
        "ash_content_treatment_sd_pct",
        "ash_content_treatment_cv_pct",
        "volatile_solids_treatment_mean_pct",
        "volatile_solids_treatment_sd_pct",
        "volatile_solids_treatment_cv_pct",
    ]

    rows_list = list(rows)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        for row in rows_list:
            formatted = {}
            for col in columns:
                value = row.get(col)
                if isinstance(value, float):
                    formatted[col] = f"{value:.6f}"
                else:
                    formatted[col] = value if value is not None else ""
            writer.writerow(formatted)


def compute_mass_loss_fresh_to_precomposted(
    treatment_rows: List[Dict[str, object]]
) -> Dict[str, object]:
    fresh_row = next(
        (
            row
            for row in treatment_rows
            if str(row.get("sample_type", "")).strip().lower() == "fresh manure"
        ),
        None,
    )
    precomp_row = next(
        (
            row
            for row in treatment_rows
            if str(row.get("sample_type", "")).strip().lower() == "precomposted manure"
        ),
        None,
    )
    if fresh_row is None or precomp_row is None:
        raise ValueError(
            "No se encontraron filas de Fresh manure y Precomposted manure en treatment_rows."
        )

    dm_fresh = float(fresh_row["dry_matter_treatment_mean_pct"])
    dm_precomp = float(precomp_row["dry_matter_treatment_mean_pct"])
    if dm_precomp == 0:
        raise ValueError("dry_matter_treatment_mean_pct de precompostado es 0.")

    ash_fresh = float(fresh_row["ash_content_treatment_mean_pct"])
    ash_precomp = float(precomp_row["ash_content_treatment_mean_pct"])
    if ash_precomp == 0:
        raise ValueError("ash_content_treatment_mean_pct de precompostado es 0.")

    dm_fresh_frac = dm_fresh / 100.0
    dm_precomp_frac = dm_precomp / 100.0
    ash_fresh_frac = ash_fresh / 100.0
    ash_precomp_frac = ash_precomp / 100.0

    # Con perdida de materia seca:
    # 1) Se estima retencion de MS por conservacion de cenizas (trazador):
    #    MS_precomp / MS_fresh = cenizas_fresco / cenizas_precomp
    dry_matter_retention_ratio = ash_fresh_frac / ash_precomp_frac

    # 2) Se convierte a retencion de masa humeda total:
    #    M_precomp / M_fresh = (MS_precomp / MS_fresh) * (DM_fresh / DM_precomp)
    mass_ratio_precomp_over_fresh = (
        dry_matter_retention_ratio * (dm_fresh_frac / dm_precomp_frac)
    )
    mass_loss_pct = (1.0 - mass_ratio_precomp_over_fresh) * 100.0

    return {
        "sampling_date": str(fresh_row.get("sampling_date", "")),
        "from_state": "Fresh manure",
        "to_state": "Precomposted manure",
        "dry_matter_from_mean_pct": dm_fresh,
        "dry_matter_to_mean_pct": dm_precomp,
        "ash_from_mean_pct_dry_basis": ash_fresh,
        "ash_to_mean_pct_dry_basis": ash_precomp,
        "dry_matter_retention_ratio": dry_matter_retention_ratio,
        "mass_ratio_to_over_from": mass_ratio_precomp_over_fresh,
        "mass_loss_pct": mass_loss_pct,
        "assumption": "same_lot_A_to_B_and_ash_mass_conservation",
        "basis": "wet_mass",
    }


def write_mass_loss_csv(row: Dict[str, object], output_path: Path) -> None:
    columns = [
        "sampling_date",
        "from_state",
        "to_state",
        "dry_matter_from_mean_pct",
        "dry_matter_to_mean_pct",
        "ash_from_mean_pct_dry_basis",
        "ash_to_mean_pct_dry_basis",
        "dry_matter_retention_ratio",
        "mass_ratio_to_over_from",
        "mass_loss_pct",
        "assumption",
        "basis",
    ]

    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=columns)
        writer.writeheader()
        formatted: Dict[str, object] = {}
        for col in columns:
            value = row.get(col)
            if isinstance(value, float):
                formatted[col] = f"{value:.6f}"
            else:
                formatted[col] = value if value is not None else ""
        writer.writerow(formatted)


# ---------------------------------------------------------------------------
# Extracción gravimétrica normalizada. Conserva las ecuaciones históricas,
# pero localiza las masas por encabezado y no consume fórmulas de resumen.
# ---------------------------------------------------------------------------

def _canonical_text(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"\s+", " ", text).strip().lower()


def _find_header_rows(worksheet) -> Tuple[int, int]:
    rows = []
    for row in worksheet.iter_rows():
        if _canonical_text(row[0].value) == "id muestra":
            rows.append(row[0].row)
    if len(rows) != 2:
        raise ValueError(f"Se esperaban dos bloques con 'ID muestra'; se encontraron {len(rows)}")
    return rows[0], rows[1]


def _header_map(worksheet, row_number: int) -> Dict[str, int]:
    return {
        _canonical_text(cell.value): cell.column
        for cell in worksheet[row_number]
        if cell.value not in (None, "")
    }


def _required_col(headers: Dict[str, int], phrase: str) -> int:
    if phrase in headers:
        return headers[phrase]
    matches = [col for header, col in headers.items() if phrase in header and "monitor" not in header]
    if len(matches) != 1:
        raise ValueError(f"Encabezado primario ambiguo o ausente: {phrase!r}; coincidencias={matches}")
    return matches[0]


def _numeric_primary(worksheet, row: int, col: int, label: str) -> float:
    value = worksheet.cell(row, col).value
    if isinstance(value, bool) or not isinstance(value, (int, float)):
        raise ValueError(f"La masa primaria {label} en {worksheet.title}!{worksheet.cell(row, col).coordinate} no es numérica")
    return float(value)


def extract_gravimetric_normalized(source: Dict[str, object], project_root: Path) -> List[Dict[str, object]]:
    """Extrae humedad, MS, cenizas y SV desde masas primarias.

    Ecuaciones preservadas:
    humedad = ((masa_húmeda - masa_seca) / masa_húmeda) * 100;
    materia seca = (masa_seca / masa_húmeda) * 100;
    cenizas = (masa_ceniza / masa_seca_incineración) * 100;
    sólidos volátiles = 100 - cenizas.
    """
    path = Path(source["absolute_path"])
    workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
    if "Data" not in workbook.sheetnames:
        raise ValueError(f"El libro no contiene la hoja Data: {path}")
    worksheet = workbook["Data"]
    drying_header_row, ash_header_row = _find_header_rows(worksheet)
    drying_headers = _header_map(worksheet, drying_header_row)
    ash_headers = _header_map(worksheet, ash_header_row)

    # Los nombres se refieren solo a masas primarias. Columnas auxiliares o de
    # resumen (incluida CRISOL ID si existe) quedan fuera de la selección.
    dry_empty = _required_col(drying_headers, "peso crisol 1 solo")
    dry_wet = _required_col(drying_headers, "peso crisol 1 con muestra humeda")
    dry_after = _required_col(drying_headers, "peso crisol 1 con muestra seca")
    ash_empty = _required_col(ash_headers, "peso crisol 2 solo")
    ash_before = _required_col(ash_headers, "peso crisol 2 con muestra seca")
    ash_after = _required_col(ash_headers, "peso crisol 2 con muestra incinerada")

    drying: Dict[str, Tuple[float, float]] = {}
    for row in range(drying_header_row + 1, ash_header_row):
        sample_id = str(worksheet.cell(row, 1).value or "").strip().upper()
        if not re.fullmatch(r"[AB]\d\d", sample_id):
            continue
        empty = _numeric_primary(worksheet, row, dry_empty, "crisol vacío de secado")
        wet_total = _numeric_primary(worksheet, row, dry_wet, "crisol con muestra húmeda")
        dry_total = _numeric_primary(worksheet, row, dry_after, "crisol con muestra seca")
        drying[sample_id] = (wet_total - empty, dry_total - empty)

    ash: Dict[str, Tuple[float, float]] = {}
    for row in range(ash_header_row + 1, worksheet.max_row + 1):
        raw_id = str(worksheet.cell(row, 1).value or "").strip().upper()
        if not re.fullmatch(r"[AB]I\d\d", raw_id):
            continue
        sample_id = raw_id.replace("I", "", 1)
        # La disposición espacial al final de algunos libros repite los IDs,
        # pero no contiene masas primarias.
        if sample_id in ash:
            continue
        empty = _numeric_primary(worksheet, row, ash_empty, "crisol vacío de incineración")
        dry_total = _numeric_primary(worksheet, row, ash_before, "crisol con muestra seca para incineración")
        ash_total = _numeric_primary(worksheet, row, ash_after, "crisol con muestra incinerada")
        ash[sample_id] = (dry_total - empty, ash_total - empty)

    if set(drying) != set(ash):
        raise ValueError(f"No coinciden los identificadores de secado e incineración: {sorted(set(drying) ^ set(ash))}")

    expected_samples = int(source["expected_samples_by_material"])
    expected_ids = {
        f"{material}{sample}{replica}"
        for material in "AB"
        for sample in range(1, expected_samples + 1)
        for replica in range(1, int(source["expected_replicates"]) + 1)
    }
    if set(drying) != expected_ids:
        raise ValueError(f"Identificadores gravimétricos inesperados en {path.name}: {sorted(set(drying) ^ expected_ids)}")

    records: List[Dict[str, object]] = []
    for sample_id in sorted(drying):
        material = "estiércol fresco" if sample_id.startswith("A") else "estiércol precompostado"
        sample_number, replica = int(sample_id[1]), int(sample_id[2])
        wet_mass, dry_mass = drying[sample_id]
        dry_inc_mass, ash_mass = ash[sample_id]
        if wet_mass == 0 or dry_inc_mass == 0:
            raise ValueError(f"Denominador cero para {sample_id}")
        values = {
            "humedad": ((wet_mass - dry_mass) / wet_mass) * 100.0,
            "materia seca": (dry_mass / wet_mass) * 100.0,
            "cenizas": (ash_mass / dry_inc_mass) * 100.0,
        }
        values["sólidos volátiles"] = 100.0 - values["cenizas"]
        normalized_id = f'{source["jornada"]}-{"EF" if material == "estiércol fresco" else "EP"}-{sample_number}'
        for variable, value in values.items():
            records.append({
                "jornada_muestreo": source["jornada"], "fecha_muestreo": source.get("sampling_date", ""),
                "fecha_recepcion": "", "fecha_analisis": "", "tipo_material": material,
                "identificador_muestra": normalized_id, "identificador_muestra_origen": sample_id,
                "repeticion_muestra": sample_number, "replica_analitica": replica,
                "nivel_observacion": "replica_analitica", "variable": variable,
                "valor": value, "unidad": "%", "base_medicion": "masa seca" if variable in {"cenizas", "sólidos volátiles"} else "masa fresca",
                "incertidumbre": "", "laboratorio": source["laboratorio"],
                "metodo_analitico": source["metodo"], "fuente_metodo_analitico": source["fuente_metodo"],
                "archivo_origen": str(path.relative_to(project_root)).replace("\\", "/"),
                "hoja_origen": "Data", "celda_o_fila_origen": f"{sample_id} / {sample_id[0]}I{sample_id[1:]}",
                "id_reporte_laboratorio": "", "uso_modelo": source["uso_modelo"],
                "motivo_uso_modelo": source["motivo_uso"], "bandera_calidad": "",
            })
    return records


def write_with_fallback(
    writer_fn, rows: Iterable[Dict[str, object]], output_path: Path
) -> Path:
    try:
        writer_fn(rows, output_path)
        return output_path
    except PermissionError:
        fallback_path = output_path.with_name(f"{output_path.stem}_updated{output_path.suffix}")
        writer_fn(rows, fallback_path)
        return fallback_path


def main() -> None:
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent
    input_path = (
        project_root
        / "Academic_documents"
        / "resultados CIA y LASA muestreo 1"
        / "Material_laboratorio_copy_to_work_python.xlsx"
    )
    output_dir = project_root / "processed"
    output_dir.mkdir(parents=True, exist_ok=True)
    detailed_output_path = output_dir / "volatile_solids_table.csv"
    representative_output_path = output_dir / "volatile_solids_representative_table.csv"
    treatment_output_path = output_dir / "volatile_solids_treatment_table.csv"
    mass_loss_output_path = output_dir / "volatile_solids_mass_loss_fresh_to_precomposted.csv"

    table = read_sheet_rows_as_table(input_path, "xl/worksheets/sheet2.xml")
    drying = extract_drying_data(table)
    incineration = extract_incineration_data(table)
    merged = merge_results(drying, incineration)
    representative_rows = summarize_representative_values(merged)
    treatment_rows = summarize_treatment_values(representative_rows)
    mass_loss_row = compute_mass_loss_fresh_to_precomposted(treatment_rows)
    final_detailed_path = write_with_fallback(write_csv, merged, detailed_output_path)
    final_representative_path = write_with_fallback(
        write_representative_csv, representative_rows, representative_output_path
    )
    final_treatment_path = write_with_fallback(
        write_treatment_csv, treatment_rows, treatment_output_path
    )
    final_mass_loss_path = write_with_fallback(
        write_mass_loss_csv, mass_loss_row, mass_loss_output_path
    )

    print(f"Generated file: {final_detailed_path}")
    print(f"Exported rows: {len(merged)}")
    print(f"Generated file: {final_representative_path}")
    print(f"Exported grouped rows: {len(representative_rows)}")
    print(f"Generated file: {final_treatment_path}")
    print(f"Exported treatment rows: {len(treatment_rows)}")
    print(f"Generated file: {final_mass_loss_path}")
    print(f"Fresh -> Precomposted mass loss (%): {mass_loss_row['mass_loss_pct']:.6f}")


if __name__ == "__main__":
    main()
